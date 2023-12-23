"""
| $Revision:: 282433                                   $:  Revision of last commit
| $Author:: wleung@SEMNET.DOM                          $:  Author of last commit
| $Date:: 2018-09-18 15:20:03 +0100 (Tue, 18 Sep 2018) $:  Date of last commit
| --------------------------------------------------------------------------------

"""
from openpyxl import worksheet
from openpyxl.styles import Border, Alignment, NamedStyle
from openpyxl import Workbook
from openpyxl.worksheet.table import Table
import openpyxl.utils
import openpyxl as xl
from openpyxl.chart import BarChart, ScatterChart, LineChart, PieChart,  Reference, Series
import glob, os


class ExcelWriter(object):
    """
    Class representing an Excel file writer which can create an excel file, and
    write data to it in a more intuitive and simplified manner.

    """
    def __init__(self):
        self.workbook = None
        self.sheet = None
        self.already_loaded = False
        self.previous_filename = None
        self.x_axis_generated = False
        
        self.first_zero = True
        self.min_index = 2
        self.max_index = 1
        self.index_col = []
        self.del_index = 2
        
    @staticmethod
    def style_range(ws, cell_range, str, style=NamedStyle()):
        """
        Apply styles to a range of cells as if they were a single cell
        https://openpyxl.readthedocs.io/en/2.4/styles.html#styling-merged-cells

        :param ws: worksheet to style
        :type ws: worksheet
        :param str: string to add to merged cells
        :type str: str
        :param style: style to apply to merged cells
        :type style: NamedStyle object
        """
        
        top = Border(top=style.border.top)
        left = Border(left=style.border.left)
        right = Border(right=style.border.right)
        bottom = Border(bottom=style.border.bottom)
        
        first_cell = ws[cell_range.split(":")[0]]
        if style.alignment:
            ws.merge_cells(cell_range)
            first_cell.alignment = style.alignment
        rows = ws[cell_range]
        if style.font:
            first_cell.font = style.font
        
        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom
        
        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = r.border + right
            if style.fill:
                for c in row:
                    c.fill = style.fill
        ws[cell_range.split(":")[0]] = str
    
    @staticmethod
    def copy_sheet(filename, wb2, sheet, delete=False):
        """
        Copies data from one sheet from a workbook to another. Does not retain formatting

        :param filename: source workbook file path
        :type filename: str
        :param wb2: destination workbook
        :type wb2: workbook
        :param sheet: sheet to copy
        :type sheet: worksheet
        :param delete: (default: False) delete sheet from source workbook
        :type delete: bool
        """
        wb1 = xl.load_workbook(filename)
        ws1 = wb1.get_sheet_by_name(sheet)
        ws2 = wb2.create_sheet(ws1.title)
        
        for row in ws1:
            for cell in row:
                ws2[cell.coordinate].value = cell.value
        
        if delete:
            wb1.remove(ws1)
            
        wb1.save(filename)
    
    @staticmethod
    def best_fit_columns(sheet, horizontal_alignment=None, skip_first_row=False):
        """
        Best fits columns to a certain extent avoiding overlapping columns

        :param sheet: sheet to edit
        :type sheet: worksheet
        :param horizontal_alignment: (default: None) option to horizontally align cells
        :type horizontal_alignment: str
        :param skip_first_row: (default: False) option to skip formatting first row
        :type skip_first_row: bool
        """
        # Value check alignment
        if horizontal_alignment and horizontal_alignment not in ['general', 'left', 'centerContinuous',
                                                                 'right', 'distributed', 'fill', 'justify', 'center']:
            raise ValueError("horizontal_alignment argument must be in {‘general’, ‘left’, ‘centerContinuous’, "
                             "‘right’, ‘distributed’, ‘fill’, ‘justify’, ‘center’}")

        # Best fitting the columns
        for col in sheet.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                if horizontal_alignment:
                    if skip_first_row:
                        if cell.row != 1:
                            alignment = Alignment(horizontal=horizontal_alignment)
                            cell.alignment = alignment
                    else:
                        alignment = Alignment(horizontal=horizontal_alignment)
                        cell.alignment = alignment
                
                # Try/Catch to avoid crashing on empty cells
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 1)
            sheet.column_dimensions[column].width = adjusted_width
     
    @staticmethod
    def create_table(data, filename, sheet_title=None, column_headings=None,
                     format_table=False):
        """
        Create a table with an option to format it like an excel table (add filters
         and a colour scheme)
         
        :param data: 2D array of strings
        :type data: list
        :param filename: file path
        :type filename: str
        :param sheet_title: (default: None) Name of sheet. Defaults to Sheet1, Sheet2, ..etc
        :type sheet_title: str
        :param column_headings: (default: None) Add column headings to an existing 2D array of data
        :type column_headings: list
        :param format_table: (default: False) Option to format as an excel table
        :type format_table: bool
        """
        # Create the workbook and change the sheet title
        workbook = Workbook()
        sheet = workbook.worksheets[0]
        sheet.title = sheet_title
        
        # Add column headings. Must be a list of strings
        if column_headings is not None:
            sheet.append(column_headings)
            
        # Append data
        for row in data:
            sheet.append(row)
        
        # Automatically create a reference to be the whole sheet
        max_column_letter = str(xl.utils.get_column_letter(sheet.max_column))
        row_count = sheet.max_row
        
        # Default Table style
        if format_table:
            medium_style = xl.worksheet.table.TableStyleInfo(name='TableStyleMedium2',
                                                             showRowStripes=True)
            tab = Table(displayName="Tab",
                        ref="A1:%s%d" % (max_column_letter, row_count),
                        tableStyleInfo=medium_style)
            
            sheet.add_table(tab)
            
        workbook.save(filename)

    @staticmethod
    def format_existing_table(sheet, first_row, vertical_headings=False):
        """
        Formats a sheet as an excel table

        :param sheet: Excel sheet to be formatted
        :type sheet: sheet
        :param vertical_headings: (default: False) Option for vertical headings to save space
        :type vertical_headings: bool
        """
        # Automatically set the table reference to be the whole sheet
        column_count = sheet.max_column
        row_count = sheet.max_row
        max_column_letter = str(xl.utils.get_column_letter(column_count))

        for column in range(column_count):
            column_letter = str(xl.utils.get_column_letter(column + 1))
            cell_value = sheet['%s%s' % (column_letter, first_row)].value
            if cell_value is None:
                cell_value = " "
            cell_value = str(cell_value)
            sheet['%s%s' % (column_letter, first_row)].value = cell_value
            if vertical_headings:
                alignment = Alignment(text_rotation='90', horizontal='center', shrink_to_fit=True)
                sheet['%s%s' % (column_letter, first_row)].alignment = alignment

        medium_style = xl.worksheet.table.TableStyleInfo(name='TableStyleMedium2',
                                                         showRowStripes=True)
        table_name = "%s" % sheet.title
        table_name.replace(" ", "")

        sheet.add_table(Table(displayName="%s" % table_name,
                              ref="A%s:%s%d" % (first_row, max_column_letter, row_count),
                              tableStyleInfo=medium_style))
        
    def create_chart(self, filename=None, sheet_name=None, chart_type='scatter',
                     independent_axis_range=None, data_range=None,
                     independent_col_name=None, data_columns_names=None,
                     col_heads_included=False, chart_title='', x_title='', y_title='',
                     index_col=None, x_log_scale=False, y_log_scale=False, split_charts=False):
        """
        Helper function that creates a chart in an excel sheet. Simplest way to use it is in the
        cleanup() function for an automated test flow.
        WARNING: Must be used right after automation because it will remove any existing charts
        
        :param filename: path to excel workbook
        :type filename: str
        :param sheet_name: name of sheet that data is collected from
        :type sheet_name: str
        :param independent_axis_range: x-axis in chart (e.g 'A1:A17')
        :type independent_axis_range: str
        :param data_range: y-axis in chart (e.g 'A1:C17, D1:D17'), column headings included
        :type data_range: str
        :param independent_col_name: name of column containing range for x-axis
        :type independent_col_name: str
        :param data_columns_names: name of column containing data for y_axis separated by commas
        :type data_columns_names: str
        :param chart_type: type of the chart (bar, pie, scatter, line)
        :type chart_type: str
        :param col_heads_included: flag if column headings included in range
        :type col_heads_included: bool
        :param chart_title: add a chart title to the chart
        :type chart_title: str
        :param x_title: add a label to the x_axis
        :type x_title: str
        :param y_title: add a label to the y_axis
        :type y_title: str
        """
        # TODO: Split charts is not working as of right now
        split_charts = False
        if filename is None:
            list_of_files = glob.glob('Data\\*.xlsx')
            filename = max(list_of_files, key=os.path.getctime)
       
        if not self.already_loaded and filename != self.previous_filename:
            self.previous_filename = filename
            self.workbook = xl.load_workbook(filename)
            self.already_loaded = True
            if sheet_name is not None:
                self.sheet = self.workbook.get_sheet_by_name(sheet_name)
            else:
                self.sheet = self.workbook.worksheets[0]
                sheet_name = self.workbook.sheetnames[0]
        else:
            self.sheet = self.workbook.worksheets[0]
            sheet_name = self.workbook.sheetnames[0]
            
        function_dict = {'bar': BarChart(), 'pie': PieChart(),
                         'scatter': ScatterChart(), 'line': LineChart()}
        
        if independent_axis_range is None and data_range is None:
            ranges = self._find_range(independent_col_name, data_columns_names, index_col)
            independent_axis_range = ranges[0]
            data_range = ranges[1]
            col_heads_included = False

        x_values = Reference(self.sheet, range_string=sheet_name + "!" + independent_axis_range)
        cs = self.workbook.create_chartsheet()
        data_range = data_range.split(',')
        data_range = [x for x in data_range if x is not '']
        if split_charts:
            for counter, data in enumerate(data_range):
                chart = function_dict[chart_type]
                
                if x_log_scale:
                    chart.x_axis.scaling.logBase = 10
                if y_log_scale:
                    chart.y_axis.scaling.logBase = 10
                    
                data = data.strip()
                values = Reference(self.sheet, range_string=sheet_name + "!" + data)
                if chart_type == 'scatter':
                    series = Series(values, x_values, title_from_data=False)
                    chart.series.append(series)
                else:
                    chart.add_data(values, titles_from_data=col_heads_included)
                    chart.set_categories(x_values)
                    
                chart.title = chart_title + 'State: %d' % counter
                chart.x_axis.title = x_title
                chart.y_axis.title = y_title
                chart.height = self.del_index
                chart.width = 20

                cs.add_chart(chart)
                
        else:
            chart = function_dict[chart_type]
            if x_log_scale:
                chart.x_axis.scaling.logBase = 10
            if y_log_scale:
                chart.y_axis.scaling.logBase = 10
    
            for data in data_range:
                data = data.strip()
                values = Reference(self.sheet, range_string=sheet_name + "!" + data)
                if chart_type == 'scatter':
                    series = Series(values, x_values, title_from_data=False)
                    chart.series.append(series)
                else:
                    chart.add_data(values, titles_from_data=col_heads_included)
                    chart.set_categories(x_values)
    
            chart.title = chart_title
            chart.x_axis.title = x_title
            chart.y_axis.title = y_title
            chart.height = 12.5
            chart.width = 20
    
            cs.add_chart(chart)
            
        self.workbook.save(filename)
        
    def _find_range(self, independent_param_name, data_params_names, index_col_name):
        max_row = self.sheet.max_row
        data_params_names = data_params_names.split(',')
        first_row = next(self.sheet.rows)
        data_range = ''
        independent_axis_range = ''
        
        if index_col_name is None:
            for index, cell in enumerate(first_row):
                if cell.value == independent_param_name:
                    independent_col = xl.utils.get_column_letter(index+1)
                    independent_axis_range = "%s2:%s%d" % (independent_col, independent_col,
                                                           max_row)
                if cell.value in data_params_names and len(data_params_names) == 1:
                    data_col = xl.utils.get_column_letter(index+1)
                    data_range += "%s1:%s%d" % (data_col, data_col, max_row)
                elif cell.value in data_params_names and len(data_params_names) > 1:
                    data_col = xl.utils.get_column_letter(index+1)
                    data_range += "%s1:%s%d," % (data_col, data_col, max_row)
        else:
            for index, cell in enumerate(first_row):
                if cell.value == index_col_name:
                    self.index_col = list(
                        next(self.sheet.iter_cols(min_col=(index + 1), max_col=(index + 1))))
                    self.index_col.pop(0)
            
            while self.max_index < self.sheet.max_row:
                self._find_indexed_range()
                for index, cell in enumerate(first_row):
                    if cell.value == independent_param_name and not self.x_axis_generated:
                        independent_col = xl.utils.get_column_letter(index + 1)
                        independent_axis_range = "%s%d:%s%d" % (independent_col,
                                                                self.min_index,
                                                                independent_col,
                                                                self.max_index)
                        self.x_axis_generated = True

                    elif cell.value in data_params_names:
                        data_col = xl.utils.get_column_letter(index+1)
                        data_range += "%s%d:%s%d," % (data_col,
                                                      self.min_index,
                                                      data_col,
                                                      self.max_index)
                self.first_zero = True
                self.min_index = self.max_index + 1

        return [independent_axis_range, data_range]
    
    def _find_indexed_range(self):
        for del_index, cell in enumerate(self.index_col):
            new_state_indicator = int(str(cell.value).split('-')[1])
            if new_state_indicator == 0 and not self.first_zero:
                self.del_index = del_index
                del self.index_col[:del_index]
                break
            else:
                self.max_index += 1
                self.first_zero = False
                
                
        

